import openpyxl
from time import clock

class Utils:
    def __init__(self,prefix):
        self.result_path = prefix[0]
        self.ext = ".xlsx"
        self.result = prefix[1]

    """
    make_excel : make dictionary data to excel file
    :param
         RESULT_PATH : 
         RESULT_EXT : .xlsx
    :return
        excel file in RESULT_PATH
    """
    def make_excel(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        sheet.cell(row=row, column=1, value="INDEX")
        sheet.cell(row=row, column=2, value="barcode")
        sheet.cell(row=row, column=3, value="guide")
        sheet.cell(row=row, column=4, value="EA")
        # for key_index, val_barcodes in sorted(self.result.items()):
        #     row = row + 1
        #     sheet.cell(row=row, column=1, value=key_index)
        #     for key_barcode, val_guides in val_barcodes.items():
        #         sheet.cell(row=row, column=2, value=key_barcode)
        #         for key_guide, val in val_guides.items():
        #             sheet.cell(row=row, column=3, value=key_guide)
        #             sheet.cell(row=row, column=4, value=len(val))
        #             j = 5
        #             for row_seq in val:
        #                 sheet.cell(row=row, column=j, value=row_seq)
        #                 j = j + 1
        #             row = row + 1
        row = 2
        for key_index, val_barcodes in sorted(self.result.items()):
            for key_barcode, val_guides in val_barcodes.items():
                for key_guide, val in val_guides.items():
                    sheet.cell(row=row, column=1, value=key_index)
                    sheet.cell(row=row, column=2, value=key_barcode)
                    sheet.cell(row=row, column=3, value=key_guide)
                    sheet.cell(row=row, column=4, value=len(val))
                    row = row + 1

        workbook.save(filename=self.result_path + str(clock()) + self.ext)